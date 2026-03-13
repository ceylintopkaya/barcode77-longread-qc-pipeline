# -*- coding: utf-8 -*-
import sys
import os
import numpy as np
from Bio import SeqIO
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Komut satiri argumanlarini al
input_file = sys.argv[1]
read_stats_output = sys.argv[2]
summary_output = sys.argv[3]
report_output = sys.argv[4]
table_output = sys.argv[5]

# Sample adini cikti dosyasinin ISMINDEN cikar (otomatik!)
SAMPLE = os.path.basename(read_stats_output).split('_read_stats.csv')[0]
if SAMPLE == "" or SAMPLE == read_stats_output:
    SAMPLE = os.path.basename(os.path.dirname(read_stats_output))

print(f"\n>>> Analiz edilen sample: {SAMPLE}")

# 1. Veriyi oku ve hesapla
results = []
for record in SeqIO.parse(input_file, "fastq"):
    length = len(record.seq)
    g = record.seq.count("G")
    c = record.seq.count("C")
    gc_content = (g + c) / length * 100 if length > 0 else 0
    quality_scores = record.letter_annotations["phred_quality"]
    mean_quality = np.mean(quality_scores)
    results.append({
        "read_id": record.id,
        "length": length,
        "gc_content": gc_content,
        "mean_quality": mean_quality
    })

df = pd.DataFrame(results)
df.to_csv(read_stats_output, index=False)

# 2. GELISMIS ISTATISTIKLER
total_reads = len(df)
total_bases = df['length'].sum()
total_yield_mb = total_bases / 1_000_000

# N50 hesapla 
sorted_lengths = sorted(df['length'].values, reverse=True)
n50 = sorted_lengths[np.argmax(np.cumsum(sorted_lengths) >= total_bases / 2)]

# Coverage (genel bir bakteri genomu ~5Mb)
genome_size_mb = 5.0
coverage = total_yield_mb / genome_size_mb

# Kalite kategorileri
q10_above = (df['mean_quality'] >= 10).sum()
q20_above = (df['mean_quality'] >= 20).sum()
q30_above = (df['mean_quality'] >= 30).sum()
q10_pct = q10_above / total_reads * 100
q20_pct = q20_above / total_reads * 100
q30_pct = q30_above / total_reads * 100

# Yazdir
print(f"\n>>> {SAMPLE} ISTATISTIKLERI:")
print(f"   Toplam reads: {total_reads:,}")
print(f"   Toplam verim: {total_yield_mb:.2f} Mb")
print(f"   Ortalama length: {df['length'].mean():.1f} bp")
print(f"   Medyan length: {df['length'].median():.1f} bp")
print(f"   N50: {n50} bp")
print(f"   Maksimum length: {df['length'].max():,.0f} bp")
print(f"   Ortalama GC: {df['gc_content'].mean():.1f}%")
print(f"   Ortalama kalite: {df['mean_quality'].mean():.1f}")

print(f"\n>>> KALITE KATEGORILERI:")
print(f"   Q10 ustu: {q10_above:,} reads ({q10_pct:.1f}%)")
print(f"   Q20 ustu: {q20_above:,} reads ({q20_pct:.1f}%)")
print(f"   Q30 ustu: {q30_above:,} reads ({q30_pct:.1f}%)")

print(f"\n>>> KAPSAMA (bakteri genomu ~5Mb):")
print(f"   Tahmini coverage: {coverage:.1f}X")

# 3. GENISLETILMIS SUMMARY TABLOSU
summary = pd.DataFrame({
    "Metric": [
        "Total_Reads", "Total_Yield_Mb", "Length_Mean_bp", 
        "Length_Median_bp", "N50_bp", "Length_Max_bp",
        "GC_Mean_Pct", "Quality_Mean",
        "Q10_Above_Pct", "Q20_Above_Pct", "Q30_Above_Pct", "Coverage_X"
    ],
    "Value": [
        f"{total_reads:,}", f"{total_yield_mb:.2f}", f"{df['length'].mean():.1f}",
        f"{df['length'].median():.1f}", f"{n50}", f"{df['length'].max():,.0f}",
        f"{df['gc_content'].mean():.1f}", f"{df['mean_quality'].mean():.1f}",
        f"{q10_pct:.1f}", f"{q20_pct:.1f}", f"{q30_pct:.1f}", f"{coverage:.1f}"
    ]
})

summary.to_csv(summary_output, index=False)
print(f">>> Summary statistics saved to {summary_output}")

# 4. TABLO GORSELI
fig, ax = plt.subplots(figsize=(10, 4))
ax.axis('off')
table_data = summary.values
table = ax.table(cellText=table_data, colLabels=summary.columns, 
                 cellLoc='center', loc='center')
table.auto_set_font_size(False)
table.set_fontsize(10)
table.scale(1.2, 1.8)
plt.title(f"{SAMPLE.capitalize()} - Summary Statistics", fontsize=14, fontweight='bold', pad=20)
plt.savefig(table_output, dpi=150, bbox_inches='tight')
print(f">>> Table saved to {table_output}")

# 5. EXCEL RAPORU (formatli)
df_clean = df.copy()
df_clean["gc_content"] = df_clean["gc_content"].round(2)
df_clean["mean_quality"] = df_clean["mean_quality"].round(2)
df_clean.columns = ["Read_ID", "Length_bp", "GC_Pct", "Quality_Q"]

with pd.ExcelWriter(report_output, engine="openpyxl") as writer:
    summary.to_excel(writer, sheet_name="Summary_Stats", index=False)
    df_clean.to_excel(writer, sheet_name="Raw_Data", index=False)

# Excel formatlamasi
wb = load_workbook(report_output)
ws = wb["Raw_Data"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

wb.save(report_output)
print(f">>> Excel report saved to {report_output}")
print(f"\n>>> TUM DOSYALAR BASARIYLA OLUSTURULDU!")